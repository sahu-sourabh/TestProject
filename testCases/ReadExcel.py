import openpyxl

wb = openpyxl.load_workbook("D:/Sourabh\PySelenium/SDET-QA_Automation_Techie/nopcommerceApp/TestData/loginData1.xlsx")
sh = wb["Sheet1"]

rows = sh.max_row
col = sh.max_column

print(rows) # 6
print(col)  # 2

for r in range(1, rows):
    for c in range(1, col+1):
        print(sh.cell(row=r, column=c).value, end="    ")
    print()
