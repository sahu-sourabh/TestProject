from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions
import openpyxl


def test_login_ddt_demo():
    # driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
    # driver.get("https://www.saucedemo.com/")

    workbook = openpyxl.load_workbook("TestData/loginData1.xlsx")
    sheet = workbook["Sheet1"]
    rows = sheet.max_row
    columns = sheet.max_column

    print("No of Rows: ", rows)
    print("No of Columns: ", columns)

    for row in range(2, rows):

        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
        driver.get("https://www.saucedemo.com/")

        username = sheet.cell(row=row, column=1).value
        password = sheet.cell(row=row, column=2).value

        driver.find_element(By.ID, "user-name").send_keys(username)
        driver.find_element(By.ID, "password").send_keys(password)
        driver.find_element(By.ID, "login-button").click()

        # time.sleep(5)

        actual_url = driver.current_url
        expected_url = "https://www.saucedemo.com/inventory.html"

        if actual_url == expected_url:
            print("********** Login Successful **********")
            WebDriverWait(driver, 10).until(
                expected_conditions.element_to_be_clickable((By.ID, "react-burger-menu-btn")))
            driver.find_element(By.ID, "react-burger-menu-btn").click()
            WebDriverWait(driver, 10).until(expected_conditions.element_to_be_clickable((By.ID, "logout_sidebar_link")))
            driver.find_element(By.ID, "logout_sidebar_link").click()
            driver.quit()
        else:
            print("********** Login Failed **********")
            driver.quit()
